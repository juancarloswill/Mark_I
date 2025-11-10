# agentes/test_y_ejecucion.py
# -*- coding: utf-8 -*-
"""
Agente de análisis, recomendación y validación (hold-out) Top-K.

Uso:
  python -m agentes.test_y_ejecucion \
    --config utils/config_optimizado.yaml \
    --backtest_xlsx outputs/backtest_plots/EURUSD_backtest_consolidado.xlsx \
    --top_k 3 \
    --outdir outputs/validacion \
    [--run_audit]
"""

from __future__ import annotations

import os
import re
import sys
import yaml
import math
import glob
import json
import argparse
import numpy as np
import pandas as pd
from pathlib import Path
from typing import Dict, List, Tuple, Optional

import logging

# -------- logging utils (no intrusivo) --------
try:
    from agentes.log_utils import setup_logging_from_config, log_cfg_snapshot, timeit, timed_block, Heartbeat
except Exception:
    # fallback mínimo si no está disponible el módulo
    def setup_logging_from_config(cfg: dict):
        logger = logging.getLogger("marki")
        if not logger.handlers:
            logger.setLevel(logging.INFO)
            ch = logging.StreamHandler()
            ch.setFormatter(logging.Formatter("%(asctime)s | %(levelname)s | %(message)s", "%Y-%m-%d %H:%M:%S"))
            logger.addHandler(ch)
        return logger
    def log_cfg_snapshot(logger, cfg, keys=None): pass
    def timeit(logger): 
        def _decor(fn):
            def _wrap(*a, **k): return fn(*a, **k)
            return _wrap
        return _decor
    def timed_block(logger, label): 
        from contextlib import contextmanager
        @contextmanager
        def _cm(): yield
        return _cm()
    class Heartbeat:
        def __init__(self, logger, total, every=10, with_memory=False): self.logger=logger; self.total=total; self.every=every; self.count=0
        def step(self, label=""): 
            self.count += 1
            if self.count % max(1, int(self.every)) == 0:
                self.logger.info(f"[HB] step={self.count} | {label}")

os.environ.setdefault("TF_CPP_MIN_LOG_LEVEL", "2")

# --------- imports dinámicos del proyecto ----------
def _import_get_model():
    for mod in ("app.utils.registry", "utils.registry", "registry", "modelos.registry"):
        try:
            m = __import__(mod, fromlist=["get_model"])
            return getattr(m, "get_model")
        except Exception:
            continue
    raise ImportError("No se pudo importar get_model desde utils/registry.py")
get_model = _import_get_model()

# === Helpers para señales/direcciones y post-proceso de CSV/Excel ===
import math
import numpy as np
import pandas as pd
from pathlib import Path

# Mapea flechas/strings a {-1,0,1}
_ARROW_MAP = {
    "↑": 1, "UP": 1, "BUY": 1, "BULL": 1, "1": 1, 1: 1,
    "→": 0, "FLAT": 0, "HOLD": 0, "0": 0, 0: 0,
    "↓": -1, "DOWN": -1, "SELL": -1, "BEAR": -1, "-1": -1, -1: -1
}
def _coerce_dir(s: pd.Series) -> pd.Series:
    s = pd.Series(s).astype(str).str.upper().str.strip()
    return s.map(_ARROW_MAP).fillna(pd.to_numeric(s, errors="coerce")).fillna(0).clip(-1,1).astype(float)

def _recompute_signal_from_price_cols(df: pd.DataFrame, pip_size: float, thr_pips: float) -> pd.Series:
    """
    Señal en precio coherente con un trade 1-step.
    Base: comparar el pronóstico contra el último observado (y_true.shift(1)).
    Fallback: si esa diferencia es 0 en toda la serie (modelo ~RW), usar delta vs y_true.
    """
    yt  = pd.to_numeric(df.get("y_true"), errors="coerce")
    yp  = pd.to_numeric(df.get("y_pred"), errors="coerce")
    ypp = pd.to_numeric(df.get("y_pred_price"), errors="coerce")
    last_price = yt.shift(1)

    # --- Base: contra último observado
    if "y_pred_price" in df.columns and ypp.notna().any():
        base_dir = np.sign(ypp - last_price)
        pips     = (ypp - last_price).abs() / float(pip_size)
    else:
        base_dir = np.sign(yp - last_price)
        pips     = (yp - last_price).abs() / float(pip_size)

    sig = pd.Series(0.0, index=df.index)
    strong = pips >= float(thr_pips)
    sig[strong & (base_dir > 0)] =  1.0
    sig[strong & (base_dir < 0)] = -1.0

    # --- Fallback: si TODOS los pips contra last_price son 0 (RW puro), usar delta vs y_true
    if pd.to_numeric(pips, errors="coerce").fillna(0).sum() == 0:
        # si hay y_pred_price, úsalo; si no, y_pred
        if "y_pred_price" in df.columns and ypp.notna().any():
            alt_dir = np.sign(ypp - yt)
            alt_pips = (ypp - yt).abs() / float(pip_size)
        else:
            alt_dir = np.sign(yp - yt)
            alt_pips = (yp - yt).abs() / float(pip_size)

        alt_sig = pd.Series(0.0, index=df.index)
        mask = alt_pips >= float(thr_pips)
        alt_sig[mask & (alt_dir > 0)] =  1.0
        alt_sig[mask & (alt_dir < 0)] = -1.0
        sig = alt_sig

    # primera fila sin referencia
    if len(sig) and (pd.isna(last_price.iloc[0]) or pd.isna(sig.iloc[0])):
        sig.iloc[0] = 0.0

    return sig.fillna(0.0)


def _ensure_direction_cols(df: pd.DataFrame) -> pd.DataFrame:
    """
    Asegura direction_true y direction_pred coherentes:
    - direction_true = sign(diff(y_true))
    - direction_pred = sign(diff(y_pred))
    """
    out = df.copy()
    if "y_true" in out.columns and "direction_true" not in out.columns:
        yt = pd.to_numeric(out["y_true"], errors="coerce")
        out["direction_true"] = np.sign(yt.diff()).fillna(0.0)
    if "y_pred" in out.columns and "direction_pred" not in out.columns:
        yp = pd.to_numeric(out["y_pred"], errors="coerce")
        out["direction_pred"] = np.sign(yp.diff()).fillna(0.0)
    return out

def _ensure_error_cols(df: pd.DataFrame) -> pd.DataFrame:
    """
    Calcula error, abs_error, sq_error si faltan (en precio).
    """
    out = df.copy()
    if {"y_true","y_pred"} <= set(out.columns):
        yt = pd.to_numeric(out["y_true"], errors="coerce")
        yp = pd.to_numeric(out["y_pred"], errors="coerce")
        if "error" not in out.columns:
            out["error"] = yp - yt
        if "abs_error" not in out.columns:
            out["abs_error"] = out["error"].abs()
        if "sq_error" not in out.columns:
            out["sq_error"] = out["error"]**2
    return out

def _fix_csvs_after_export(outdir: Path, pip_size: float, thr_pips: float) -> None:
    """
    Post-proceso no intrusivo sobre los CSV:
    - Normaliza nombres/tipos.
    - Asegura direction_* y error_*.
    - Recalcula 'signal' con fallback RW; si queda toda en 0, conserva la original si es mejor.
    - Si direction_pred queda toda en 0, usa delta vs y_true como último recurso.
    """
    for p in sorted(outdir.glob("*.csv")):
        try:
            df = pd.read_csv(p)
        except Exception:
            continue

        ren = {"ytrue":"y_true", "ypred":"y_pred", "pred":"y_pred",
               "price_pred":"y_pred_price", "Signal":"signal",
               "Direction_Pred":"direction_pred", "Direction_True":"direction_true"}
        df = df.rename(columns={c: ren.get(c, ren.get(str(c).lower(), c)) for c in df.columns})

        # ordenar por tiempo si viene 'ds'
        if "ds" in df.columns:
            df["ds"] = pd.to_datetime(df["ds"], errors="coerce")
            df = df.sort_values("ds").reset_index(drop=True)

        # numeric & flechas
        for c in ("y_true","y_pred","y_pred_price","signal","direction_pred","direction_true"):
            if c in df.columns:
                if c.startswith("direction"):
                    s = pd.Series(df[c]).astype(str).str.upper().str.strip()
                    df[c] = s.map(_ARROW_MAP).fillna(pd.to_numeric(s, errors="coerce")).fillna(0).clip(-1,1)
                else:
                    df[c] = pd.to_numeric(df[c], errors="coerce")

        # asegurar columnas
        df = _ensure_direction_cols(df)
        df = _ensure_error_cols(df)

        # señal nueva
        sig_new = _recompute_signal_from_price_cols(df, pip_size=pip_size, thr_pips=thr_pips)
        if "signal" in df.columns:
            # si la nueva queda toda en 0, pero la original tiene algo de señal, conservar la original
            if pd.to_numeric(sig_new, errors="coerce").fillna(0).abs().sum() == 0:
                sig_old = pd.to_numeric(df["signal"], errors="coerce").fillna(0)
                df["signal"] = sig_old
            else:
                df["signal"] = sig_new
        else:
            df["signal"] = sig_new

        # direction_pred: si queda todo 0, intenta con y_pred_price/y_pred vs y_true
        if "direction_pred" in df.columns and pd.to_numeric(df["direction_pred"], errors="coerce").fillna(0).abs().sum() == 0:
            yt  = pd.to_numeric(df.get("y_true"), errors="coerce")
            yp  = pd.to_numeric(df.get("y_pred"), errors="coerce")
            ypp = pd.to_numeric(df.get("y_pred_price"), errors="coerce")
            last_price = yt.shift(1)
            if "y_pred_price" in df.columns and ypp.notna().any():
                df["direction_pred"] = np.sign(ypp - last_price).fillna(0.0)
            else:
                df["direction_pred"] = np.sign(yp - last_price).fillna(0.0)

        df.to_csv(p, index=False)


def _overwrite_excel_signals(
    excel_path: Path,
    pred_map_valid: dict[str, pd.DataFrame],
    pip_size: float,
    thr_pips: float,
    start_row: int = 13,
    header_row: int = 12,
    header_scan_rows: int = 30  # NUEVO: escaneo para autodetectar encabezado
) -> None:
    """
    Reescribe 'signal' y 'direction_pred' por encabezado detectado.
    - Autodetecta fila de encabezado si 'header_row' no coincide.
    - Si la señal/dirección recalculadas quedan TODAS en 0/NaN, usa las del df original.
    - Repara y_true en la hoja solo si está todo 0/NaN.
    """
    from openpyxl import load_workbook

    if not excel_path.exists():
        return

    # alias de encabezados (minúsculas y sin espacios/acentos)
    alias = {
        "y_true": {"y_true","ytrue","true","y","real","precio_real","y true","y-true"},
        "y_pred": {"y_pred","ypred","pred","forecast","yhat","y pred","y-pred"},
        "y_pred_price": {"y_pred_price","price_pred","ypred_price","yhat_price","y pred price","y-pred-price"},
        "direction_pred": {"direction_pred","dir_pred","d_pred","pred_dir","direction pred","direction-pred","pred direction"},
        "direction_true": {"direction_true","dir_true","d_true","true_dir","direction true","direction-true","true direction"},
        "signal": {"signal","señal","senial"}
    }

    def _norm(s: str | None) -> str:
        return "" if s is None else str(s).strip().lower()

    def _find_header_row(ws) -> tuple[dict[str,int], int] | tuple[None, None]:
        """Devuelve (mapa_encabezados, fila_encabezado) o (None,None) si no encuentra."""
        # intenta primero la fila provista
        candidates = [header_row] + list(range(1, min(header_scan_rows, ws.max_row) + 1))
        seen = set()
        for r in candidates:
            if r < 1 or r > ws.max_row or r in seen:
                continue
            seen.add(r)
            try:
                row_vals = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
            except Exception:
                continue
            raw_map = {}
            for idx, v in enumerate(row_vals, start=1):
                k = _norm(v)
                if k:
                    raw_map[k] = idx
            # consideramos que es encabezado si al menos 3 columnas clave se detectan
            keys_found = sum(1 for k in ("y_pred","y_true","signal","direction_pred") if k in raw_map)
            if keys_found >= 2:  # relajado
                return raw_map, r
        return None, None

    # carga libro
    book = load_workbook(excel_path)

    # mapeo de nombres de hoja tolerante a mayúsculas/minúsculas
    normalized_sheetnames = {s.lower(): s for s in book.sheetnames}

    for sheet_name, dfp in (pred_map_valid or {}).items():
        if sheet_name not in book.sheetnames:
            # intenta variantes
            alt = normalized_sheetnames.get(sheet_name.lower())
            if not alt:
                continue
            sheet_name = alt

        ws = book[sheet_name]

        # detectar encabezado
        raw_map, found_header_row = _find_header_row(ws)
        if not raw_map:
            continue

        def _find_col(key: str) -> int | None:
            k = _norm(key)
            if k in raw_map:
                return raw_map[k]
            for a in alias.get(k, {k}):
                a_n = _norm(a)
                if a_n in raw_map:
                    return raw_map[a_n]
            return None

        # ordenar/indizar el DF del modelo
        tmp = dfp.copy()
        if isinstance(tmp.index, pd.DatetimeIndex):
            tmp = tmp.sort_index()
        elif "ds" in tmp.columns:
            tmp["ds"] = pd.to_datetime(tmp["ds"], errors="coerce")
            tmp = tmp.sort_values("ds").set_index("ds")

        max_rows = ws.max_row
        n_lim = max(0, min(len(tmp), max_rows - (start_row - 1)))
        if n_lim <= 0:
            continue

        # -------------------------------
        # 1) SIGNAL (con fallback seguro)
        # -------------------------------
        col_signal = _find_col("signal")
        if col_signal:
            sig_calc = _recompute_signal_from_price_cols(tmp, pip_size=pip_size, thr_pips=thr_pips).astype(float)
            # fallback: si todo 0/NaN, intenta usar la que viene del dfp (si existe y no es toda 0)
            use_series = sig_calc
            if pd.to_numeric(sig_calc, errors="coerce").fillna(0).abs().sum() == 0 and "signal" in tmp.columns:
                sig_orig = pd.to_numeric(tmp["signal"], errors="coerce").fillna(0.0)
                if sig_orig.abs().sum() > 0:
                    use_series = sig_orig
            vals = use_series.iloc[:n_lim].tolist()
            for i, v in enumerate(vals, start=start_row):
                ws.cell(row=i, column=col_signal, value=float(v))
            for i in range(start_row + n_lim, ws.max_row + 1):
                ws.cell(row=i, column=col_signal, value=None)

        # -----------------------------------------
        # 2) DIRECTION_PRED (con fallback seguro)
        # -----------------------------------------
        col_dirp = _find_col("direction_pred")
        if col_dirp:
            yt  = pd.to_numeric(tmp.get("y_true"), errors="coerce")
            yp  = pd.to_numeric(tmp.get("y_pred"), errors="coerce")
            ypp = pd.to_numeric(tmp.get("y_pred_price"), errors="coerce")
            last_price = yt.shift(1)
            if "y_pred_price" in tmp.columns and ypp.notna().any():
                d_calc = np.sign(ypp - last_price)
            else:
                d_calc = np.sign(yp - last_price)
            d_calc = pd.Series(d_calc, index=tmp.index).fillna(0.0).astype(float)

            use_dir = d_calc
            if use_dir.abs().sum() == 0 and "direction_pred" in tmp.columns:
                d_orig = pd.to_numeric(tmp["direction_pred"], errors="coerce").fillna(0.0)
                if d_orig.abs().sum() > 0:
                    use_dir = d_orig.astype(float)

            vals = use_dir.iloc[:n_lim].tolist()
            for i, v in enumerate(vals, start=start_row):
                ws.cell(row=i, column=col_dirp, value=float(v))
            for i in range(start_row + n_lim, ws.max_row + 1):
                ws.cell(row=i, column=col_dirp, value=None)

        # -----------------------------------------
        # 3) Y_TRUE: repara solo si la hoja está en 0
        # -----------------------------------------
        col_ytrue = _find_col("y_true")
        if col_ytrue and ("y_true" in tmp.columns):
            # mirar tramo visible
            excel_vals = []
            for i in range(start_row, start_row + n_lim):
                excel_vals.append(ws.cell(row=i, column=col_ytrue).value)
            s_excel = pd.to_numeric(pd.Series(excel_vals), errors="coerce")
            if s_excel.fillna(0).abs().sum() == 0:  # todo 0/NaN en la hoja
                yt = pd.to_numeric(tmp["y_true"], errors="coerce").fillna(0.0)
                for i, v in enumerate(yt.iloc[:n_lim].tolist(), start=start_row):
                    ws.cell(row=i, column=col_ytrue, value=float(v))

    book.save(excel_path)



def _import_exporters():
    for mod in ("reportes.reportes_excel", "app.reportes.reportes_excel", "reportes_excel"):
        try:
            m = __import__(mod, fromlist=[
                "export_backtest_csv_per_model",
                "export_backtest_excel_consolidado"
            ])
            return (
                getattr(m, "export_backtest_csv_per_model"),
                getattr(m, "export_backtest_excel_consolidado"),
            )
        except Exception:
            continue
    raise ImportError("No se encontraron exportadores en reportes/reportes_excel.py")
export_backtest_csv_per_model, export_backtest_excel_consolidado = _import_exporters()

try:
    from app.eda.eda_crispdm import _ensure_dt_index, _find_close, _resample_ohlc  # type: ignore
except Exception:
    def _ensure_dt_index(df: pd.DataFrame) -> pd.DataFrame:
        return df.sort_index()
    def _find_close(df: pd.DataFrame) -> str:
        return "Close" if "Close" in df.columns else df.columns[-1]
    def _resample_ohlc(df: pd.DataFrame, freq: str = "D", price_col: str = "Close") -> pd.DataFrame:
        return df

_HAS_MT5 = True
try:
    from conexion.easy_Trading import Basic_funcs as _BF
except Exception as _e:
    _HAS_MT5 = False
    _BF = None
    print(f"⚠️ No se pudo importar Basic_funcs (conexion.easy_Trading): {_e}")

# --------- utils ----------
# === Señales y direcciones (numéricas) ===
import numpy as np
import pandas as pd

def _calc_directions(y: pd.Series) -> pd.Series:
    """
    Dirección como sign(delta): -1 (baja), 0 (plana), 1 (sube).
    Primer valor queda NaN y luego se rellena con 0.
    """
    d = pd.Series(y, copy=True).astype(float).diff()
    out = np.sign(d)
    return out.fillna(0.0)

def _calc_signal(
    y_true: pd.Series,
    y_pred: pd.Series,
    pip_size: float,
    min_threshold_pips: float
) -> pd.Series:
    """
    Señal de trading basada en el cambio PREDICTO respecto al último verdadero.
    - delta = y_pred - y_true.shift(1)
    - umbral = min_threshold_pips * pip_size
    - signal = 1 si delta > +umbral; -1 si delta < -umbral; 0 en otro caso.
    """
    yt = pd.Series(y_true, copy=True).astype(float)
    yp = pd.Series(y_pred, copy=True).astype(float)

    thr = float(min_threshold_pips) * float(pip_size)
    delta = yp - yt.shift(1)

    sig = np.where(delta >  thr,  1.0,
          np.where(delta < -thr, -1.0, 0.0))

    # Primera fila suele no tener referencia -> pon 0
    if len(sig) > 0 and np.isnan(delta.iloc[0]):
        sig[0] = 0.0

    return pd.Series(sig, index=yp.index)


def _load_config(require_path: str) -> dict:
    p = Path(require_path)
    if not p.is_file():
        raise FileNotFoundError(f"No se encontró el archivo de configuración: {p}")
    with open(p, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)

def _apply_data_window(df: pd.DataFrame, cfg: dict) -> pd.DataFrame:
    if not isinstance(cfg, dict) or not cfg:
        return df
    mode = str(cfg.get("mode", "last_n_bars")).lower()
    if mode == "date_range":
        start = cfg.get("start", None)
        end   = cfg.get("end", None)
        if start is not None: start = pd.to_datetime(start)
        if end   is not None: end   = pd.to_datetime(end)
        if start is None and end is None:
            return df
        if start is not None and end is not None:
            return df.loc[start:end]
        elif start is not None:
            return df.loc[start:]
        else:
            return df.loc[:end]
    else:
        n = int(cfg.get("n_bars", 0))
        if 0 < n < len(df):
            return df.iloc[-n:]
        return df

@timeit(logging.getLogger("marki"))
def _split_train_valid(price: pd.Series, cfg_valid: dict) -> Tuple[pd.Series, Optional[pd.Series]]:
    if not isinstance(cfg_valid, dict) or not cfg_valid:
        return price, None
    modo = str(cfg_valid.get("modo", "last_n")).lower()
    if modo == "date_range":
        start = cfg_valid.get("start", None)
        end   = cfg_valid.get("end", None)
        if start is not None: start = pd.to_datetime(start)
        if end   is not None: end   = pd.to_datetime(end)
        if start is None and end is None:
            return price, None
        mask_valid = pd.Series(True, index=price.index)
        if start is not None: mask_valid &= (price.index >= start)
        if end   is not None: mask_valid &= (price.index <= end)
    else:
        n = int(cfg_valid.get("n", 0))
        if n <= 0 or n >= len(price):
            return price, None
        mask_valid = pd.Series(False, index=price.index)
        mask_valid.iloc[-n:] = True
    price_valid = price[mask_valid]
    price_train = price[~mask_valid]
    if len(price_train) < 50 or len(price_valid) < 10:
        return price, None
    return price_train, price_valid

@timeit(logging.getLogger("marki"))
def _get_series_from_mt5(config: dict) -> Tuple[pd.DataFrame, str]:
    if not _HAS_MT5:
        raise RuntimeError("MT5 no disponible: no es posible reconstruir la serie.")
    simbolo   = config.get("simbolo", "EURUSD")
    timeframe = config.get("timeframe", "D1")
    cantidad  = int(config.get("cantidad_datos", 1500))
    mt5c = config.get("mt5", {})
    bf = _BF(mt5c.get("login"), mt5c.get("password"), mt5c.get("server"), mt5c.get("path"))  # type: ignore
    print("✅ Conexión establecida con MetaTrader 5 (test_y_ejecucion)")
    df = bf.get_data_for_bt(timeframe, simbolo, cantidad)
    ren = {"open":"Open","high":"High","low":"Low","close":"Close","tick_volume":"TickVolume","real_volume":"Volume","time":"Date"}
    for k,v in ren.items():
        if k in df.columns and v not in df.columns:
            df = df.rename(columns={k:v})
    if "Date" in df.columns:
        df = df.set_index("Date")
    df = df.sort_index()
    price_col = _find_close(df)
    df = _ensure_dt_index(df)
    freq_cfg = str(config.get("eda", {}).get("frecuencia_resampleo", "D")).upper()
    with timed_block(logging.getLogger("marki"), "resample_ohlc"):
        df = _resample_ohlc(df, freq=freq_cfg, price_col=price_col)
    df = _apply_data_window(df, config.get("data_window", {}) or {})
    return df, price_col

# --------- leer XLSX backtest ----------
@timeit(logging.getLogger("marki"))
def _read_backtest_consolidado(xlsx_path: str) -> Tuple[pd.DataFrame, Dict[str, pd.DataFrame]]:
    from openpyxl import load_workbook
    wb = load_workbook(filename=xlsx_path, data_only=True, read_only=True)
    sheets = wb.sheetnames
    if "metrics" not in sheets:
        raise ValueError("El archivo no contiene hoja 'metrics'.")
    ws = wb["metrics"]
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(h) if h is not None else "" for h in header]
    metrics = pd.DataFrame(list(ws.iter_rows(min_row=2, values_only=True)), columns=headers)
    per_model: Dict[str, pd.DataFrame] = {}
    for name in sheets:
        if name.lower() in {"metrics", "config_info"}:
            continue
        ws2 = wb[name]
        header2 = next(ws2.iter_rows(min_row=1, max_row=1, values_only=True))
        headers2 = [str(h) if h is not None else "" for h in header2]
        dfm = pd.DataFrame(list(ws2.iter_rows(min_row=2, values_only=True)), columns=headers2)
        if "ds" in dfm.columns:
            dfm["ds"] = pd.to_datetime(dfm["ds"], errors="coerce")
            dfm = dfm.set_index("ds")
            dfm.index.name = "ds"
        per_model[name] = dfm
    return metrics, per_model

# --------- selección TOP-K (HR -> RMSE -> DM simple) ----------
@timeit(logging.getLogger("marki"))
def _pick_topk_models(metrics: pd.DataFrame,
                      per_model: Dict[str, pd.DataFrame],
                      k: int,
                      prefer_col_hr: List[str] = ["Direction_Accuracy","Hit Rate","Hit_Rate","HitRate","DA","DA%"],
                      prefer_col_rmse: List[str] = ["RMSE","rmse"]) -> List[str]:
    if metrics is None or metrics.empty:
        raise ValueError("No hay métricas en 'metrics'.")

    cols = {c.lower(): c for c in metrics.columns}

    hr_col = None
    for c in prefer_col_hr:
        if c in metrics.columns: hr_col = c; break
        if c.lower() in cols:    hr_col = cols[c.lower()]; break
    if hr_col is None:
        raise ValueError("No se encontró columna de Hit Rate / Direction_Accuracy.")

    rmse_col = None
    for c in prefer_col_rmse:
        if c in metrics.columns: rmse_col = c; break
        if c.lower() in cols:    rmse_col = cols[c.lower()]; break
    if rmse_col is None:
        raise ValueError("No se encontró columna RMSE.")

    name_col = None
    for c in ("model","Modelo","name","Name","MODEL"):
        if c in metrics.columns:
            name_col = c; break
    if name_col is None:
        cand = [c for c in metrics.columns if metrics[c].dtype==object]
        name_col = cand[0] if cand else metrics.columns[0]

    df_sorted = metrics.sort_values(by=[hr_col, rmse_col], ascending=[False, True]).reset_index(drop=True)
    topk = df_sorted.head(int(k)).copy()

    def _get_errors(sheet: str) -> Optional[pd.Series]:
        if sheet not in per_model:
            return None
        dfm = per_model[sheet].copy()
        y_true = None; y_pred = None
        for yt in ("y_true","true","Y"):
            if yt in dfm.columns:
                y_true = pd.to_numeric(dfm[yt], errors="coerce"); break
        for yp in ("y_pred","yhat","pred","forecast"):
            if yp in dfm.columns:
                y_pred = pd.to_numeric(dfm[yp], errors="coerce"); break
        if y_true is None or y_pred is None:
            return None
        y_true, y_pred = y_true.align(y_pred, join="inner")
        return (y_true - y_pred).dropna()

    def dm_stat(e1: pd.Series, e2: pd.Series, h: int = 1, power: int = 2) -> float:
        e1, e2 = e1.align(e2, join="inner")
        if len(e1) < 20:
            return 0.0
        d = (e1.abs() ** power) - (e2.abs() ** power)
        d = d - d.mean()
        lag = max(h - 1, 0)
        gamma0 = float((d * d).mean())
        s = gamma0
        for l in range(1, lag + 1):
            w = 1.0 - l / (lag + 1.0)
            gamma = float((d[l:] * d[:-l]).mean())
            s += 2.0 * w * gamma
        return float(d.mean() / math.sqrt(s / len(d))) if s > 0 else 0.0

    i = 0
    chosen = []
    while i < len(topk):
        row = topk.iloc[i]
        same = topk[(topk[hr_col]==row[hr_col]) & (topk[rmse_col]==row[rmse_col])]
        if len(same) == 1:
            chosen.append(str(row[name_col])); i += 1; continue
        names = [str(r[name_col]) for _, r in same.iterrows()]
        errors_map = {n: _get_errors(n) for n in names}
        if any(v is None or v.empty for v in errors_map.values()):
            chosen.extend(names); i += len(same); continue
        mse_avg = {n: float((errors_map[n]**2).mean()) for n in names}
        names_sorted = sorted(names, key=lambda n: mse_avg[n])
        anchor = names_sorted[0]
        out_order = [anchor]
        for cand in names_sorted[1:]:
            _ = abs(dm_stat(errors_map[anchor], errors_map[cand], h=1, power=2))
            out_order.append(cand)
        chosen.extend(out_order); i += len(same)
    return chosen[:int(k)]

# --------- señales y metrica auxiliar ----------
def _compute_signal(yhat: float, last_price: float, config: dict, target: str) -> int:
    bt = (config.get("bt") or {})
    pip_size = float(bt.get("pip_size", 0.0001))
    thr_pips = float(bt.get("threshold_pips", 12.0))
    if target == "returns":
        thr_ret = (thr_pips * pip_size) / max(abs(last_price), 1e-9)
        if yhat >  thr_ret:  return 1
        if yhat < -thr_ret:  return -1
        return 0
    else:
        thr_abs = thr_pips * pip_size
        if (yhat - last_price) >  thr_abs: return 1
        if (yhat - last_price) < -thr_abs: return -1
        return 0

def _fixed_window_1d(series: pd.Series, win: int) -> np.ndarray:
    x = series.astype("float32").to_numpy()
    if len(x) >= win:
        tail = x[-win:]
    else:
        pad = np.full((win - len(x),), x[0] if len(x)>0 else 0.0, dtype="float32")
        tail = np.concatenate([pad, x.astype("float32")], axis=0)
    return tail.astype("float32")

# --------- validación rolling 1-paso ----------
@timeit(logging.getLogger("marki"))
def _roll_validate_on_test(model_name: str,
                           model_params: dict,
                           price_train: pd.Series,
                           price_test: pd.Series,
                           config: dict) -> pd.DataFrame:
    logger = logging.getLogger("marki")
    target = str((config.get("bt", {}) or {}).get("target", "returns")).lower()
    freq   = str((config.get("eda", {}) or {}).get("frecuencia_resampleo", "D"))
    cfg_local = {"target": target, "freq": "H" if freq.upper().startswith("H") else "D"}

    key = model_name.strip().lower()
    cfg_local[key] = model_params or {}

    stride = int((config.get("agent", {}) or {}).get("validation", {}).get("refit_stride", 1))
    stride = max(1, stride)

    logger.info(f"[valid] model={model_name} target={target} stride={stride} params={model_params}")

    model = get_model(key, cfg_local)
    cur   = price_train.copy()

    # --- Intento de preentrenado (LSTM), sin romper flujo existente ---
    skip_initial_fit = False
    if key == "lstm" and bool((model_params or {}).get("reuse_pretrained", False)):
        if hasattr(model, "load_pretrained") and model.load_pretrained():
            logger.info("LSTM: modelo+scaler preentrenados cargados. Se salta fit inicial.")
            skip_initial_fit = True
        else:
            logger.info("LSTM: no se encontró preentrenado util; se entrenará desde cero.")

    # FIT inicial (solo si no cargamos preentrenado)
    if not skip_initial_fit:
        with timed_block(logger, f"fit_inicial:{model_name}"):
            model.fit(cur)

    # Heartbeat para loops largos
    profile_cfg = ((config.get("logging") or {}).get("profile") or {})
    hb = Heartbeat(
        logger=logger,
        total=len(price_test.index),
        every=int(profile_cfg.get("every_n_steps", 10) or 10),
        with_memory=bool(profile_cfg.get("with_memory", False))
    )

    # ---- LSTM: warm-up para “congelar” traza (evitar retracing) ----
    lstm_warm_done = False

    preds = []
    step  = 0
    for ts in price_test.index:
        step += 1

        # (A) update/refit antes de predecir
        if step > 1:
            updated = False
            for meth in ("update", "partial_fit", "fit_partial"):
                if hasattr(model, meth):
                    try:
                        with timed_block(logger, f"update:{model_name}"):
                            getattr(model, meth)(cur)
                        updated = True
                        break
                    except Exception:
                        pass
            force_classic_refit = (key in ("arima", "sarima", "prophet", "rw"))
            if force_classic_refit or (not updated and (step - 1) % stride == 0):
                try:
                    with timed_block(logger, f"refit:{model_name}"):
                        model.fit(cur)
                except Exception:
                    pass

        # (B) predecir 1-paso
        with timed_block(logger, f"predict:{model_name}"):
            if key == "lstm":
                win = int((model_params or {}).get("window", 64))
                w1d = _fixed_window_1d(cur, win)

                # warm-up (una sola vez)
                if not lstm_warm_done:
                    try:
                        _ = model.predict(1, last_timestamp=cur.index[-1], last_window=w1d)
                    except TypeError:
                        _ = model.predict(1)
                    lstm_warm_done = True

                try:
                    dfp = model.predict(1, last_timestamp=cur.index[-1], last_window=w1d)
                    if not isinstance(dfp, pd.DataFrame):
                        dfp = pd.DataFrame({"y_pred": [float(np.ravel(dfp)[0])]}, index=pd.DatetimeIndex([ts]))
                except TypeError:
                    dfp = model.predict(1)
            else:
                try:
                    dfp = model.predict(h=1, last_timestamp=cur.index[-1])
                except TypeError:
                    dfp = model.predict(1)

        if "yhat" in getattr(dfp, "columns", []):
            dfp = dfp.rename(columns={"yhat":"y_pred"})
        dfp.index = pd.DatetimeIndex([ts])

        # (C) conversión returns/price coherente
        last_price = float(cur.iloc[-1])
        yhat = float(dfp["y_pred"].iloc[0])
        if target == "returns":
            # modelos base devuelven nivel => a retorno aproximado
            if key in ("arima","sarima","prophet","rw","lstm"):
                yhat = (yhat / max(last_price, 1e-9)) - 1.0
                dfp.loc[dfp.index, "y_pred"] = yhat

        # señal y precio predicho
        signal = _compute_signal(yhat, last_price, config, target)
        dfp["signal"] = int(signal)
        dfp["y_pred_price"] = last_price * (1.0 + yhat) if target == "returns" else float(yhat)

        # ventana usada
        try:
            dfp["fecha_inicio_ventana"] = cur.index.min()
            dfp["fecha_fin_ventana"] = cur.index.max()
        except Exception:
            pass
        
        # y_true del timestamp actual (si existe en price_test)
        try:
            true_val = price_test.loc[[ts]].iloc[0]  # <-- necesitas price_test visible o pásalo por cierre
        except Exception:
            true_val = np.nan
        dfp["y_true"] = true_val
        
        # calcula direcciones de 1 paso (respecto al último precio observado)
        if target == "returns":
            dir_pred = float(np.sign(yhat))                               # retorno predicho
            dir_true = float(np.sign((true_val / max(last_price, 1e-9)) - 1.0)) if pd.notna(true_val) else 0.0
        else:
            dir_pred = float(np.sign(float(dfp["y_pred"].iloc[0]) - last_price))
            dir_true = float(np.sign(true_val - last_price)) if pd.notna(true_val) else 0.0

        dfp["direction_pred"] = dir_pred
        dfp["direction_true"] = dir_true

        preds.append(dfp[[
            "y_pred", "y_pred_price", "y_true", "signal",
            "direction_pred", "direction_true",
            "fecha_inicio_ventana", "fecha_fin_ventana"
        ]])

        # (D) revelar obs real y extender ventana
        cur = pd.concat([cur, price_test.loc[[ts]]])

        # heartbeat
        hb.step(label=f"{model_name} rolling")

    out = pd.concat(preds).sort_index()
    out.index.name = "ds"
    return out

# --------- export espejo backtest ----------
@timeit(logging.getLogger("marki"))
def _export_validation(symbol: str,
                       pred_map_valid: Dict[str, pd.DataFrame],
                       price_valid: pd.Series,
                       config: dict,
                       outdir: Path,
                       excel_path: Path) -> None:
    outdir.mkdir(parents=True, exist_ok=True)
    export_backtest_csv_per_model(
        symbol=symbol,
        pred_map=pred_map_valid,
        price=price_valid,
        outdir=outdir,
        target=str((config.get("bt",{}) or {}).get("target","returns")).lower(),
        pip_size=float((config.get("bt",{}) or {}).get("pip_size", 0.0001)),
        threshold_mode=str((config.get("bt",{}) or {}).get("threshold_mode","fixed")).lower(),
        threshold_pips=float((config.get("bt",{}) or {}).get("threshold_pips", 12.0)),
        horizon=1,
        initial_train=None,
    )
    export_backtest_excel_consolidado(
        symbol=symbol,
        pred_map=pred_map_valid,
        price=price_valid,
        excel_path=excel_path,
        target=str((config.get("bt",{}) or {}).get("target","returns")).lower(),
        pip_size=float((config.get("bt",{}) or {}).get("pip_size", 0.0001)),
        threshold_mode=str((config.get("bt",{}) or {}).get("threshold_mode","fixed")).lower(),
        threshold_pips=float((config.get("bt",{}) or {}).get("threshold_pips", 12.0)),
        horizon=1,
        annualization=None,
        per_model_params=None,
        config_info={"validacion": config.get("validacion", {})},
        seasonality_m=1,
        initial_train=None,
    )

    # --- Post-proceso no intrusivo: arregla CSVs y sobrescribe señales en Excel ---
    bt_cfg = (config.get("bt") or {})
    _pip = float(bt_cfg.get("pip_size", 0.0001))
    _thr = float(bt_cfg.get("threshold_pips", 12.0))

    # 1) corrige CSVs exportados (añade direction_*, error_* y recalcula signal)
    _fix_csvs_after_export(outdir=outdir, pip_size=_pip, thr_pips=_thr)

    # 2) sobreescribe columna C (signal) desde fila 13 en cada pestaña de modelo
    _overwrite_excel_signals(excel_path=excel_path,
                             pred_map_valid=pred_map_valid,
                             pip_size=_pip, thr_pips=_thr,
                             start_row=13)

    logging.getLogger("marki").info(f"💾 Validación (hold-out) exportada en {excel_path}")

def _write_extra_sheets(backtest_xlsx: str, validation_xlsx: Path, metrics_valid_df: pd.DataFrame) -> None:
    logger = logging.getLogger("marki")
    cfg_df = None
    try:
        xls = pd.ExcelFile(backtest_xlsx)
        if 'config_info' in xls.sheet_names:
            cfg_df = pd.read_excel(xls, 'config_info')
    except Exception as e:
        logger.info(f"ℹ️ No fue posible leer 'config_info' del backtest: {e}")

    try:
        with pd.ExcelWriter(validation_xlsx, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            if metrics_valid_df is not None and not metrics_valid_df.empty:
                metrics_valid_df.to_excel(writer, sheet_name='metrics_valid', index=False)
            if cfg_df is not None:
                cfg_df.to_excel(writer, sheet_name='config_info', index=False)
        logger.info(f"💾 Hojas adicionales escritas en {validation_xlsx} (metrics_valid, config_info)")
    except FileNotFoundError:
        with pd.ExcelWriter(validation_xlsx, engine='openpyxl', mode='w') as writer:
            if metrics_valid_df is not None and not metrics_valid_df.empty:
                metrics_valid_df.to_excel(writer, sheet_name='metrics_valid', index=False)
            if cfg_df is not None:
                cfg_df.to_excel(writer, sheet_name='config_info', index=False)
        logger.info(f"💾 Archivo de validación creado y hojas escritas en {validation_xlsx}")

# --------- métricas valid ----------
def _series_pred(df_pred: pd.DataFrame) -> pd.Series:
    tmp = df_pred.copy()
    if not isinstance(tmp.index, pd.DatetimeIndex) and "ds" in tmp.columns:
        tmp = tmp.set_index("ds")
    if "y_pred" in tmp.columns:
        return tmp["y_pred"].astype(float)
    elif "yhat" in tmp.columns:
        return tmp["yhat"].astype(float)
    for c in reversed(list(tmp.columns)):
        s = pd.to_numeric(tmp[c], errors="coerce")
        if s.notna().sum() > 0:
            return s
    raise ValueError("No se encontró columna de predicción en DF.")

def _dm_pvalue_test(y_true: pd.Series, yhat_a: pd.Series, yhat_b: pd.Series) -> float | None:
    y_true, yhat_a = y_true.align(yhat_a, join="inner")
    y_true, yhat_b = y_true.align(yhat_b, join="inner")
    yhat_a = yhat_a.reindex(y_true.index).astype(float)
    yhat_b = yhat_b.reindex(y_true.index).astype(float)
    if len(y_true) < 20:
        return None
    e1 = (y_true - yhat_a).astype(float)
    e2 = (y_true - yhat_b).astype(float)
    d = (e1.abs()**2) - (e2.abs()**2)
    d = d - d.mean()
    gamma0 = float((d * d).mean())
    if gamma0 <= 0: return None
    dm_stat = float(d.mean()) / math.sqrt(gamma0 / len(d))
    from math import erf, sqrt
    def _phi_cdf(z): return 0.5 * (1 + erf(z / sqrt(2)))
    p = 2.0 * (1.0 - _phi_cdf(abs(dm_stat)))
    return p

@timeit(logging.getLogger("marki"))
def _metrics_from_pred_map(pred_map_valid: Dict[str, pd.DataFrame], price_valid: pd.Series) -> pd.DataFrame:
    """
    Calcula métricas por modelo usando:
    - Direction_Accuracy:
        1) Si existen direction_pred/true y NO son todos cero, se usan.
        2) Si direction_pred existe pero es todo 0/NaN -> se recalcula como sign(y_pred_price - y_{t-1})
           (o sign(y_pred - y_{t-1}) si no hay y_pred_price).
        3) Si no existen columnas de dirección -> se infiere:
            - returns: sign(y_pred) vs sign(%Δ y_true)
            - price:   sign(y_pred - y_{t-1}) vs sign(y_true - y_{t-1})
    - MAE/RMSE/MAPE en precio (si y_pred está en retornos, se convierte a precio 1-paso).
    """
    rows = []
    for name, dfp in (pred_map_valid or {}).items():
        tmp = dfp.copy()
        if not isinstance(tmp.index, pd.DatetimeIndex) and 'ds' in tmp.columns:
            tmp['ds'] = pd.to_datetime(tmp['ds'], errors='coerce')
            tmp = tmp.set_index('ds')

        # Serie base del bloque de validación
        y_true = price_valid.copy().astype(float).reindex(tmp.index)

        # y_pred (tal como venga)
        if 'y_pred' in tmp.columns:
            y_pred = pd.to_numeric(tmp['y_pred'], errors='coerce')
        elif 'yhat' in tmp.columns:
            y_pred = pd.to_numeric(tmp['yhat'], errors='coerce')
        else:
            num_cols = [c for c in tmp.columns if pd.api.types.is_numeric_dtype(tmp[c])]
            y_pred = pd.to_numeric(tmp[num_cols[-1]], errors='coerce') if num_cols else pd.Series(index=tmp.index, dtype=float)

        # ---------- Direction_Accuracy con fallback robusto ----------
        da = np.nan

        def _rebuild_dirs_like_excel(df_):
            yt  = pd.to_numeric(df_.get("y_true"), errors="coerce")
            yp  = pd.to_numeric(df_.get("y_pred"), errors="coerce")
            ypp = pd.to_numeric(df_.get("y_pred_price"), errors="coerce")
            last = yt.shift(1)
            if "y_pred_price" in df_.columns and ypp.notna().any():
                d_pred = np.sign(ypp - last)
            else:
                d_pred = np.sign(yp - last)
            d_true = np.sign(yt - last)
            return pd.Series(d_pred, index=df_.index).fillna(0.0), pd.Series(d_true, index=df_.index).fillna(0.0)

        if {'direction_pred', 'direction_true'} <= set(tmp.columns):
            dp = pd.to_numeric(tmp['direction_pred'], errors='coerce').clip(-1,1)
            dt = pd.to_numeric(tmp['direction_true'], errors='coerce').clip(-1,1)
            # Si direction_pred quedó toda 0/NaN (caso ARIMA retornos ~0), rehacer como en el Excel/CSV
            if dp.fillna(0).abs().sum() == 0:
                dp, dt = _rebuild_dirs_like_excel(tmp.assign(y_true=y_true, y_pred=y_pred))
            m = (~dp.isna()) & (~dt.isna())
            da = float((dp[m] == dt[m]).mean() * 100.0) if m.any() else np.nan
        else:
            # Inferir escala para construir DA
            med_pred = float(pd.Series(y_pred).abs().median(skipna=True) or 0.0)
            med_true = float(pd.Series(y_true).abs().median(skipna=True) or 0.0)
            looks_like_returns = med_pred < 0.2 and med_true > 0.5
            if looks_like_returns:
                rt_true = (y_true / y_true.shift(1) - 1.0)
                rt_pred = y_pred
                m = (~rt_true.isna()) & (~rt_pred.isna())
                da = float((np.sign(rt_true[m]) == np.sign(rt_pred[m])).mean() * 100.0) if m.any() else np.nan
            else:
                last = y_true.shift(1)
                m = (~last.isna()) & (~y_true.isna()) & (~y_pred.isna())
                da = float((np.sign(y_pred[m] - last[m]) == np.sign(y_true[m] - last[m])).mean() * 100.0) if m.any() else np.nan

        # ---------- Errores (en PRICE) ----------
        if 'y_pred_price' in tmp.columns:
            ypp = pd.to_numeric(tmp['y_pred_price'], errors='coerce').reindex(y_true.index)
            yp_price = ypp
        else:
            looks_like_returns = (float(pd.Series(y_pred).abs().median(skipna=True) or 0.0) < 0.2
                                  and float(pd.Series(y_true).abs().median(skipna=True) or 0.0) > 0.5)
            if looks_like_returns:
                yp_price = y_true.shift(1) * (1.0 + y_pred)
            else:
                yp_price = y_pred

        yt, yp = y_true.align(yp_price, join='inner')
        err = (yt - yp).astype(float)
        mae  = float(err.abs().mean())   if len(err) else float('nan')
        rmse = float(np.sqrt((err**2).mean())) if len(err) else float('nan')
        with np.errstate(divide='ignore', invalid='ignore'):
            mape = float(np.nanmean(np.abs(err) / np.where(np.abs(yt) < 1e-12, np.nan, np.abs(yt))) * 100.0) if len(err) else float('nan')

        rows.append({'model': str(name), 'MAE': mae, 'RMSE': rmse, 'MAPE': mape, 'Direction_Accuracy': da})

    dfm = pd.DataFrame(rows)
    if not dfm.empty:
        dfm = dfm.sort_values(by=['Direction_Accuracy','RMSE'], ascending=[False, True]).reset_index(drop=True)
        dfm.insert(0, 'rank', range(1, len(dfm)+1))
    return dfm

# ============================================================
# =======  BLOQUE DE AUDITORÍA (opcional, no intrusivo)  =====
# ============================================================

_POSSIBLE_SEPS = [",", ";", "\t", "|"]
_KEYVAL_PATTERN = re.compile(r"([A-Za-z0-9_]+)\s*=\s*([^,\s\|;]+)")

def _discover_validacion_csvs(base_dir: str | Path) -> List[str]:
    """Busca CSVs en la carpeta de validación (ARIMA/LSTM/PROPHET/RW primero)."""
    base = Path(base_dir).expanduser().resolve()
    if not base.exists():
        return []
    patrones = [
        str(base / "*ARIMA*.csv"),
        str(base / "*LSTM*.csv"),
        str(base / "*PROPHET*.csv"),
        str(base / "*RW*.csv"),
        str(base / "*.csv"),
    ]
    vistos, hallados = set(), []
    for pat in patrones:
        for p in glob.glob(pat):
            ap = str(Path(p).resolve())
            if ap not in vistos:
                vistos.add(ap)
                hallados.append(ap)
    return hallados

def _normalize_common(df: pd.DataFrame) -> pd.DataFrame:
    """Normaliza nombres y tipos más comunes (y_true/y_pred/etc.)."""
    rename_map = {
        "ds":"ds", "date":"ds", "timestamp":"ds", "time":"ds",
        "ytrue":"y_true", "yTrue":"y_true",
        "ypred":"y_pred", "yPred":"y_pred",
        "pred":"y_pred", "price_pred":"y_pred_price",
    }
    for c in list(df.columns):
        cc = c.strip()
        if cc in rename_map:
            df = df.rename(columns={c: rename_map[cc]})
        elif cc.lower() in rename_map:
            df = df.rename(columns={c: rename_map[cc.lower()]})
    # index temporal si existe
    for c in ["ds","date","timestamp","time"]:
        if c in df.columns:
            try:
                df[c] = pd.to_datetime(df[c], errors="coerce")
                df = df.set_index(c)
                break
            except Exception:
                pass
    # coerción numérica de columnas de interés
    for c in ["y_true","y_pred","y_pred_price","error","abs_error","sq_error","signal","direction_pred","direction_true"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")
    for c in ["direction_pred","direction_true"]:
        if c in df.columns:
            df[c] = _coerce_dir(df[c])  # ← mapea ↑/↓/BUY/SELL/1/0/-1 a {-1,0,1}
    return df

def _try_sep_read(path: str) -> pd.DataFrame | None:
    for sep in _POSSIBLE_SEPS:
        try:
            df = pd.read_csv(path, sep=sep)
            if df.shape[1] > 1:
                return df
        except Exception:
            pass
    return None

def _try_json_per_line(path: str) -> pd.DataFrame | None:
    rows = []
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                obj = json.loads(line)
                if isinstance(obj, dict):
                    rows.append(obj)
            except Exception:
                try:
                    obj = json.loads(line.replace("'", '"'))
                    if isinstance(obj, dict):
                        rows.append(obj)
                except Exception:
                    continue
    if rows:
        return pd.DataFrame(rows)
    return None

def _try_keyval_parse(path: str) -> pd.DataFrame | None:
    rows = []
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            kv = dict(_KEYVAL_PATTERN.findall(line))
            if kv:
                rows.append(kv)
    if rows:
        return pd.DataFrame(rows)
    return None

def load_backtest_csv(path: str) -> pd.DataFrame | None:
    """Carga robusta de CSVs (separadores, JSON por línea o pares clave=valor)."""
    if not os.path.isfile(path):
        return None
    # 1) separadores
    df = _try_sep_read(path)
    if df is not None and df.shape[1] > 1:
        return _normalize_common(df)
    # 2) json per line
    df = _try_json_per_line(path)
    if df is not None:
        return _normalize_common(df)
    # 3) key=value
    df = _try_keyval_parse(path)
    if df is not None:
        return _normalize_common(df)
    # 4) fallback extremo: partir una columna por separadores
    try:
        df_raw = pd.read_csv(path, header=None, names=["raw"])
        for sep in _POSSIBLE_SEPS:
            parts = df_raw["raw"].str.split(sep, expand=True)
            if parts.shape[1] >= 2:
                df_try = parts.copy()
                header_row = df_try.iloc[0].astype(str).str.strip().tolist()
                if all(h and h.replace(" ","").isalpha() for h in header_row):
                    df_try.columns = header_row
                    df_try = df_try.iloc[1:].reset_index(drop=True)
                return _normalize_common(df_try)
    except Exception:
        pass
    return None

def infer_scale(series: pd.Series) -> str:
    s = pd.Series(series).dropna().astype(float)
    if len(s) == 0:
        return "unknown"
    med = float(s.abs().median())
    maxv = float(s.abs().max())
    if med < 0.01 and maxv < 0.2:
        return "returns"
    if 0.5 <= float(s.mean()) <= 5.0:
        return "price"
    return "unknown"

def basic_metrics(y_true: pd.Series, y_pred: pd.Series) -> Dict[str, float]:
    yt = pd.Series(y_true).astype(float).to_numpy()
    yp = pd.Series(y_pred).astype(float).to_numpy()
    m = ~np.isnan(yt) & ~np.isnan(yp)
    yt = yt[m]; yp = yp[m]
    if len(yt) == 0:
        return {"n": 0, "MAE": np.nan, "RMSE": np.nan, "MAPE": np.nan}
    err = yp - yt
    mae = float(np.mean(np.abs(err)))
    rmse = float(np.sqrt(np.mean(err**2)))
    with np.errstate(divide="ignore", invalid="ignore"):
        mape = float(np.nanmean(np.abs(err) / np.where(np.abs(yt) < 1e-12, np.nan, np.abs(yt))) * 100.0)
    return {"n": int(len(yt)), "MAE": mae, "RMSE": rmse, "MAPE": mape}

def describe_signals(sig: pd.Series) -> Dict[str, float]:
    s = pd.Series(sig).dropna().astype(float)
    n = len(s)
    if n == 0:
        return {"n": 0, "p_buy": np.nan, "p_hold": np.nan, "p_sell": np.nan}
    return {
        "n": n,
        "p_buy": float((s == 1).mean() * 100.0),
        "p_hold": float((s == 0).mean() * 100.0),
        "p_sell": float((s == -1).mean() * 100.0)
    }

def _run_audit_on_outdir(outdir: Path, pip_size: float = 0.0001, out_xlsx: Optional[Path] = None) -> None:
    """
    Audita los CSV exportados en `outdir` y escribe:
    - backtest_ranges: métricas por archivo
    - signals_distribution: distribución de señales
    - dm_tests: comparación simple entre modelos (si procede)
    """
    logger = logging.getLogger("marki")
    csvs = _discover_validacion_csvs(outdir)
    if not csvs:
        logger.info("ℹ️ Auditoría: no se encontraron CSVs en la carpeta de validación.")
        return

    bt_rows, sig_rows = [], []
    model_errors: Dict[str, pd.Series] = {}

    for path in csvs:
        name = os.path.basename(path)
        df = load_backtest_csv(path)
        if df is None or df.empty:
            continue

        cols = set(df.columns)
        row = {"file": name, "n_rows": int(len(df))}
        ycols = [c for c in ["y_true", "y_pred", "y_pred_price"] if c in cols]
        scols = [c for c in ["signal", "direction_pred", "direction_true"] if c in cols]
        row["y_cols"] = ",".join(ycols)
        row["signal_cols"] = ",".join(scols)

        if "y_true" in df.columns and "y_pred" in df.columns:
            mets = basic_metrics(df["y_true"], df["y_pred"])
            row.update({f"pred_{k}": v for k, v in mets.items()})
            row["scale_true"] = infer_scale(df["y_true"])
            row["scale_pred"] = infer_scale(df["y_pred"])
            # para DM
            yt, yp = df["y_true"].align(df["y_pred"], join="inner")
            e = (yt - yp).dropna()
            if len(e) > 10:
                model_errors[os.path.splitext(name)[0]] = e

        if "y_true" in df.columns and "y_pred_price" in df.columns:
            mets_p = basic_metrics(df["y_true"], df["y_pred_price"])
            row.update({f"pred_price_{k}": v for k, v in mets_p.items()})

        bt_rows.append(row)

        for sc in scols:
            d = describe_signals(df[sc])
            sig_rows.append({"file": name, "signal_col": sc, **d})

    backtests_df = pd.DataFrame(bt_rows)
    if not backtests_df.empty and "file" in backtests_df.columns:
        backtests_df = backtests_df.sort_values(["file"]).reset_index(drop=True)

    signals_df = pd.DataFrame(sig_rows)
    if not signals_df.empty and {"file","signal_col"}.issubset(signals_df.columns):
        signals_df = signals_df.sort_values(["file", "signal_col"]).reset_index(drop=True)

    # DM simple entre errores (normal aprox.)
    dm_out = []
    keys = list(model_errors.keys())
    for i in range(len(keys)):
        for j in range(i + 1, len(keys)):
            a, b = keys[i], keys[j]
            e1, e2 = model_errors[a].align(model_errors[b], join="inner")
            if len(e1) < 20:
                continue
            d = (e1.abs()**2) - (e2.abs()**2)
            d = d - d.mean()
            gamma0 = float((d * d).mean())
            if gamma0 <= 0:
                continue
            dm_stat = float(d.mean()) / math.sqrt(gamma0 / len(d))
            # p-valor normal aprox
            from math import erf, sqrt
            def _phi_cdf(z): return 0.5 * (1 + erf(z / sqrt(2)))
            p = 2.0 * (1.0 - _phi_cdf(abs(dm_stat)))
            dm_out.append({"A": a, "B": b, "DM_stat": dm_stat, "p_value": p})

    dm_df = pd.DataFrame(dm_out)

    # escribir excel de auditoría
    if out_xlsx is None:
        out_xlsx = outdir / "auditoria_validacion.xlsx"
    with pd.ExcelWriter(out_xlsx, engine="xlsxwriter") as wr:
        if not backtests_df.empty:
            backtests_df.to_excel(wr, index=False, sheet_name="backtest_ranges")
        if not signals_df.empty:
            signals_df.to_excel(wr, index=False, sheet_name="signals_distribution")
        if not dm_df.empty:
            dm_df.to_excel(wr, index=False, sheet_name="dm_tests")
    logger.info(f"✅ Auditoría escrita en {out_xlsx}")
    
# ===================== AUX: señales + métricas + resumen =====================

import os, re, json
import numpy as np
import pandas as pd
from pathlib import Path
from typing import Dict, List, Optional, Tuple

# Mapea flechas/strings a valores -1/0/1
_ARROW_MAP = {
    "↑": 1, "UP": 1, "BUY": 1, "BULL": 1, "1": 1, 1: 1,
    "→": 0, "FLAT": 0, "HOLD": 0, "0": 0, 0: 0,
    "↓": -1, "DOWN": -1, "SELL": -1, "BEAR": -1, "-1": -1, -1: -1
}

def _coerce_dir(s: pd.Series) -> pd.Series:
    s = s.astype(str).str.upper().str.strip()
    return s.map(_ARROW_MAP).fillna(pd.to_numeric(s, errors="coerce")).fillna(0).astype(float)

def recompute_signal(df: pd.DataFrame, pip_size: float, thr_pips: float) -> pd.Series:
    """
    Devuelve -1/0/+1 aplicando un umbral de pips a la magnitud de la predicción
    y usando la dirección del propio modelo.
    """
    # 1) dirección base
    if "direction_pred" in df.columns:
        base_dir = _coerce_dir(df["direction_pred"]).clip(-1, 1)
    else:
        # fallback: dirección por diferencia de y_pred vs y_true (simétrico)
        base_dir = np.sign(pd.to_numeric(df.get("y_pred", 0), errors="coerce") -
                           pd.to_numeric(df.get("y_true", 0), errors="coerce"))
        base_dir = pd.Series(base_dir, index=df.index).fillna(0).clip(-1, 1)

    # 2) magnitud en pips (ideal: y_pred_price vs last_price; fallback: y_pred vs y_true)
    if {"y_pred", "y_true"}.issubset(df.columns):
        yp = pd.to_numeric(df["y_pred"], errors="coerce")
        yt = pd.to_numeric(df["y_true"], errors="coerce")
        pips = ((yp - yt).abs() / float(pip_size))
        strong = pips >= float(thr_pips)
        sig = pd.Series(0.0, index=df.index)
        sig[strong & (base_dir > 0)] = 1.0
        sig[strong & (base_dir < 0)] = -1.0
        return sig.fillna(0.0)
    else:
        return base_dir.fillna(0.0)

def direction_accuracy(df: pd.DataFrame) -> float:
    """% de acierto direccional usando direction_pred vs direction_true."""
    if {"direction_pred", "direction_true"}.issubset(df.columns):
        a = _coerce_dir(df["direction_pred"])
        b = _coerce_dir(df["direction_true"])
        m = (~a.isna()) & (~b.isna())
        return float((a[m] == b[m]).mean() * 100.0) if m.any() else float("nan")
    return float("nan")

def basic_metrics(y_true: pd.Series, y_pred: pd.Series) -> Dict[str, float]:
    yt = pd.to_numeric(y_true, errors="coerce").to_numpy()
    yp = pd.to_numeric(y_pred, errors="coerce").to_numpy()
    m = ~np.isnan(yt) & ~np.isnan(yp)
    if not m.any():
        return {"n": 0, "MAE": np.nan, "RMSE": np.nan, "MAPE": np.nan}
    e = yp[m] - yt[m]
    mae = float(np.mean(np.abs(e)))
    rmse = float(np.sqrt(np.mean(e**2)))
    with np.errstate(divide="ignore", invalid="ignore"):
        mape = float(np.nanmean(np.abs(e) / np.where(np.abs(yt[m]) < 1e-12, np.nan, np.abs(yt[m]))) * 100.0)
    return {"n": int(m.sum()), "MAE": mae, "RMSE": rmse, "MAPE": mape}

def _try_read_csv(path: str) -> Optional[pd.DataFrame]:
    """Lector robusto: maneja separadores comunes o una línea JSON/key=value."""
    seps = [",", ";", "\t", "|"]
    for sep in seps:
        try:
            df = pd.read_csv(path, sep=sep)
            if df.shape[1] > 1:
                return df
        except Exception:
            pass
    # 1-col -> intenta explotar
    try:
        raw = pd.read_csv(path, header=None, names=["raw"])
        for sep in seps:
            parts = raw["raw"].str.split(sep, expand=True)
            if parts.shape[1] >= 2:
                header = parts.iloc[0].astype(str).str.strip().tolist()
                # si parece header, úsalo
                if all(h and h.replace(" ", "") for h in header):
                    parts.columns = header
                    return parts.iloc[1:].reset_index(drop=True)
                return parts
    except Exception:
        pass
    return None

def _normalize_cols(df: pd.DataFrame) -> pd.DataFrame:
    ren = {
        "ytrue": "y_true", "yTrue": "y_true", "y_pred_price":"y_pred_price",
        "ypred": "y_pred", "yPred": "y_pred", "pred": "y_pred",
        "ds":"ds", "date":"ds", "timestamp":"ds", "time":"ds"
    }
    # normaliza nombres
    newcols = {}
    for c in df.columns:
        lc = c.strip()
        newcols[c] = ren.get(lc, ren.get(lc.lower(), c))
    df = df.rename(columns=newcols)
    # numéricos clave
    for c in ["y_true","y_pred","y_pred_price","signal","direction_pred","direction_true"]:
        if c in df.columns:
            if c.startswith("direction"):
                df[c] = _coerce_dir(df[c])
            else:
                df[c] = pd.to_numeric(df[c], errors="coerce")
    return df

def fix_csvs_and_compute_summary(outdir: str, pip_size: float, thr_pips: float) -> Dict[str, Dict[str, float]]:
    """
    1) Relee todos los CSV de validación en 'outdir'
    2) Recalcula 'signal'
    3) Calcula RMSE/MAE/MAPE y Direction Accuracy
    4) Devuelve métricas por modelo (dict)
    """
    out = {}
    base = Path(outdir)
    csvs = sorted([p for p in base.glob("*.csv") if p.suffix.lower()==".csv"])
    for p in csvs:
        df = _try_read_csv(str(p))
        if df is None or df.empty:
            continue
        df = _normalize_cols(df)

        # Señal recalculada (sobrescribe)
        df["signal"] = recompute_signal(df, pip_size=pip_size, thr_pips=thr_pips)

        # Métricas
        mets = {"DA": direction_accuracy(df)}
        if {"y_true","y_pred"}.issubset(df.columns):
            mets.update(basic_metrics(df["y_true"], df["y_pred"]))

        # Guarda CSV corregido
        df.to_csv(p, index=False)

        model_name = p.stem  # ej: EURUSD_LSTM_backtest
        out[model_name] = mets
    return out

# ---- Diebold–Mariano sobre errores de precio (opcional) ----
from scipy.stats import t as student_t

def _align(a: pd.Series, b: pd.Series) -> Tuple[pd.Series, pd.Series]:
    idx = a.dropna().index.intersection(b.dropna().index)
    return a.loc[idx], b.loc[idx]

def dm_test_from_csvs(outdir: str, a_key: str, b_key: str) -> Dict[str, float]:
    """Calcula DM entre modelos (por nombre de archivo sin .csv, p.ej. 'EURUSD_ARIMA_backtest')."""
    A = Path(outdir, f"{a_key}.csv")
    B = Path(outdir, f"{b_key}.csv")
    if not A.exists() or not B.exists():
        return {"DM_stat": np.nan, "p_value": np.nan}
    da = _normalize_cols(_try_read_csv(str(A)) or pd.DataFrame())
    db = _normalize_cols(_try_read_csv(str(B)) or pd.DataFrame())
    if not {"y_true","y_pred"}.issubset(da.columns) or not {"y_true","y_pred"}.issubset(db.columns):
        return {"DM_stat": np.nan, "p_value": np.nan}
    ea = pd.to_numeric(da["y_true"], errors="coerce") - pd.to_numeric(da["y_pred"], errors="coerce")
    eb = pd.to_numeric(db["y_true"], errors="coerce") - pd.to_numeric(db["y_pred"], errors="coerce")
    ea, eb = _align(ea, eb)
    if len(ea) < 10 or len(eb) < 10:
        return {"DM_stat": np.nan, "p_value": np.nan}
    d = (ea**2 - eb**2).dropna()
    if len(d) < 10:
        return {"DM_stat": np.nan, "p_value": np.nan}
    dbar = d.mean()
    T = len(d)
    # var(d) (sin corrección HAC para h>1; aquí h=1)
    var_d = ((d - dbar)**2).sum() / T
    if var_d <= 0:
        return {"DM_stat": np.nan, "p_value": np.nan}
    DM = float(dbar / np.sqrt(var_d / T))
    p = float(2.0 * (1.0 - student_t.cdf(abs(DM), df=max(T-1, 1))))
    return {"DM_stat": DM, "p_value": p}

# =========================
# --------- main ----------
# =========================
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--config", required=True, help="Ruta a config.yaml (obligatoria)")
    ap.add_argument("--backtest_xlsx", required=True, help="Ruta al XLSX consolidado del backtest")
    ap.add_argument("--outdir", default="outputs/validacion", help="Carpeta de salida para validación")
    ap.add_argument("--top_k", type=int, default=1, help="Numero de mejores modelos a validar (TOP-K)")
    ap.add_argument("--run_audit", action="store_true", help="Si se indica, corre auditoría sobre los CSV exportados")
    args = ap.parse_args()

    config = _load_config(args.config)

    # === Logger desde YAML ===
    logger = setup_logging_from_config(config)
    log_cfg_snapshot(logger, config)

    logger.info("Leyendo backtest consolidado…")
    metrics, per_model = _read_backtest_consolidado(args.backtest_xlsx)

    logger.info("Seleccionando Top-K modelos…")
    topk_names = _pick_topk_models(metrics, per_model, k=max(1, args.top_k))
    logger.info(f"⭐ TOP-{len(topk_names)} modelos (HR→RMSE→DM): {topk_names}")

    logger.info("Descargando serie MT5 y preparando split…")
    df, price_col = _get_series_from_mt5(config)
    price = df[price_col].astype(float)
    price_train, price_test = _split_train_valid(price, config.get("validacion", {}) or {})
    if price_test is None:
        logger.warning("⚠️ No hay bloque de test definido en config['validacion']. Se detiene la validación.")
        sys.exit(0)
    logger.info(f"Split -> train={len(price_train)}, test={len(price_test)} | fechas: train[{price_train.index.min()}..{price_train.index.max()}] test[{price_test.index.min()}..{price_test.index.max()}]")

    name2params_cfg = {str(m.get("name","")).strip().lower(): (m.get("params", {}) or {}) for m in (config.get("modelos") or [])}

    pred_map_valid: Dict[str, pd.DataFrame] = {}
    for name_raw in topk_names:
        key = name_raw.strip().lower()
        params_cfg = name2params_cfg.get(key, {})
        logger.info(f"⚙️ {name_raw}: usando parámetros de config -> {params_cfg}")
        pred_valid = _roll_validate_on_test(name_raw, params_cfg, price_train, price_test, config)
        pred_map_valid[name_raw] = pred_valid
        logger.info(f"[VALID] Terminado {name_raw}: {len(pred_valid)} filas")

    symbol = re.sub(r"[^A-Za-z0-9_]+","_", str(config.get("simbolo","SYMB")).upper())
    outdir = Path(args.outdir)
    outdir.mkdir(parents=True, exist_ok=True)
    excel_path = outdir / "validacion_consolidado.xlsx"

    _export_validation(symbol, pred_map_valid, price_test, config, outdir, excel_path)

    metrics_valid = _metrics_from_pred_map(pred_map_valid, price_test)
    if len(pred_map_valid) >= 1 and not metrics_valid.empty and "model" in metrics_valid.columns:
        top1 = list(pred_map_valid.keys())[0]
        yhat_top1 = _series_pred(pred_map_valid[top1])
        dm_pvals = {}
        for name, dfp in pred_map_valid.items():
            if name == top1:
                dm_pvals[name] = 0.0
            else:
                p = _dm_pvalue_test(price_test, _series_pred(dfp), yhat_top1)
                dm_pvals[name] = float(p) if p is not None else np.nan
        metrics_valid["DM_pvalue"] = metrics_valid["model"].map(dm_pvals)

    _write_extra_sheets(args.backtest_xlsx, excel_path, metrics_valid)
    logger.info("✅ Proceso de validación finalizado.")

    # -------- Auditoría opcional sobre los CSV recién exportados --------
    if args.run_audit:
        bt_cfg = (config.get("bt") or {})
        pip_size = float(bt_cfg.get("pip_size", 0.0001))
        _run_audit_on_outdir(outdir=outdir, pip_size=pip_size, out_xlsx=outdir / "auditoria_validacion.xlsx")

# --- RESUMEN EJECUTIVO (no intrusivo) ---
    try:
        valid_xlsx_path = os.path.join(args.outdir, "validacion_consolidado.xlsx")
        write_summary_text(valid_xlsx_path, args.outdir)
    except Exception as e:
        print(f"[WARN] Resumen ejecutivo no generado: {e}")



# =================== RESUMEN EJECUTIVO (no intrusivo) ===================

import os, glob, re, json
import numpy as np
import pandas as pd
from pathlib import Path

_POSSIBLE_SEPS = [",", ";", "\t", "|"]
_KEYVAL_PATTERN = re.compile(r"([A-Za-z0-9_]+)\s*=\s*([^,\s\|;]+)")

def _try_sep_read(path: str) -> pd.DataFrame | None:
    for sep in _POSSIBLE_SEPS:
        try:
            df = pd.read_csv(path, sep=sep)
            if df.shape[1] > 1:
                return df
        except Exception:
            pass
    return None

def _try_json_per_line(path: str) -> pd.DataFrame | None:
    rows = []
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                obj = json.loads(line)
                if isinstance(obj, dict):
                    rows.append(obj)
            except Exception:
                try:
                    obj = json.loads(line.replace("'", '"'))
                    if isinstance(obj, dict):
                        rows.append(obj)
                except Exception:
                    continue
    if rows:
        return pd.DataFrame(rows)
    return None

def _try_keyval_parse(path: str) -> pd.DataFrame | None:
    rows = []
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            kv = dict(_KEYVAL_PATTERN.findall(line))
            if kv:
                rows.append(kv)
    if rows:
        return pd.DataFrame(rows)
    return None

def _normalize_common(df: pd.DataFrame) -> pd.DataFrame:
    # normaliza nombres comunes y tipos
    rename_map = {
        "ds":"ds", "date":"ds", "timestamp":"ds", "time":"ds",
        "ytrue":"y_true", "yTrue":"y_true",
        "ypred":"y_pred", "yPred":"y_pred",
        "pred":"y_pred", "price_pred":"y_pred_price",
        "direction_pred":"direction_pred",
        "direction_true":"direction_true",
        "signal":"signal"
    }
    for c in list(df.columns):
        cc = str(c).strip()
        if cc in rename_map:
            df = df.rename(columns={c: rename_map[cc]})
        elif cc.lower() in rename_map:
            df = df.rename(columns={c: rename_map[cc.lower()]})
    # index temporal si existe
    for c in ["ds","date","timestamp","time"]:
        if c in df.columns:
            try:
                df[c] = pd.to_datetime(df[c], errors="coerce")
                df = df.set_index(c)
                break
            except Exception:
                pass
    # coerción numérica
    for c in ["y_true","y_pred","y_pred_price","error","abs_error","sq_error","signal","direction_pred","direction_true"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")
    return df

def _load_validacion_csvs(validacion_dir: str | Path) -> dict[str, pd.DataFrame]:
    base = Path(validacion_dir).expanduser().resolve()
    out: dict[str, pd.DataFrame] = {}
    patrones = ["*ARIMA*backtest*.csv", "*LSTM*backtest*.csv", "*PROPHET*backtest*.csv", "*RW*backtest*.csv"]
    files = []
    for pat in patrones:
        files.extend(glob.glob(str(base / pat)))
    # fallback: cualquier csv
    if not files:
        files = glob.glob(str(base / "*.csv"))
    for p in files:
        name = Path(p).stem  # p.ej. EURUSD_ARIMA_backtest
        # carga robusta
        df = _try_sep_read(p)
        if df is None:
            df = _try_json_per_line(p)
        if df is None:
            df = _try_keyval_parse(p)
        if df is None:
            try:
                raw = pd.read_csv(p, header=None, names=["raw"])
                # explotar por separador si se detecta
                best = None
                for sep in _POSSIBLE_SEPS:
                    parts = raw["raw"].str.split(sep, expand=True)
                    if parts.shape[1] >= 2:
                        best = parts
                        break
                df = best if best is not None else raw
            except Exception:
                continue
        df = _normalize_common(df)
        out[name] = df
    return out

def _direction_accuracy(df: pd.DataFrame) -> float | None:
    """
    Calcula acierto direccional:
    - Si existen 'direction_pred' y 'direction_true' en {-1,0,1}, usa coincidencia exacta.
    - Si no, usa signo de diferencias: sign(y_pred - y_{t-1}) vs sign(y_true - y_{t-1}) si es posible,
      o sign(y_pred - y_true) como último recurso (menos ideal).
    Devuelve porcentaje (0-100).
    """
    try:
        s = df.dropna()
        if "direction_pred" in s.columns and "direction_true" in s.columns:
            d = s[["direction_pred","direction_true"]].dropna()
            if len(d) == 0:
                return None
            return float((d["direction_pred"] == d["direction_true"]).mean() * 100.0)
        # fallback: signos por primera diferencia
        if "y_true" in s.columns and "y_pred" in s.columns:
            y_true = s["y_true"].astype(float)
            y_pred = s["y_pred"].astype(float)
            # intentamos comparar contra y_{t-1} (requiere índice ordenado)
            if y_true.index.is_monotonic_increasing and len(y_true) > 1:
                yt_diff = y_true.diff()
                yp_diff = y_pred.diff()
                m = (~yt_diff.isna()) & (~yp_diff.isna())
                if m.sum() > 0:
                    return float((np.sign(yt_diff[m]) == np.sign(yp_diff[m])).mean() * 100.0)
            # última opción: coincidencia de signo del error (no ideal, pero informativa)
            err = y_pred - y_true
            return float((np.sign(err) == 0).mean() * 100.0)  # casi siempre ~0
    except Exception:
        return None
    return None

def _rmse(y_true: pd.Series, y_pred: pd.Series) -> float | None:
    try:
        yt = pd.to_numeric(y_true, errors="coerce").to_numpy()
        yp = pd.to_numeric(y_pred, errors="coerce").to_numpy()
        m = ~np.isnan(yt) & ~np.isnan(yp)
        if m.sum() == 0:
            return None
        return float(np.sqrt(np.mean((yt[m] - yp[m])**2)))
    except Exception:
        return None

def _align(a: pd.Series, b: pd.Series) -> tuple[pd.Series, pd.Series]:
    idx = a.dropna().index.intersection(b.dropna().index)
    return a.loc[idx], b.loc[idx]

from scipy.stats import t as student_t

def _dm_test(ea: pd.Series, eb: pd.Series, h: int = 1, power: int = 2) -> tuple[float | None, float | None]:
    # DM clásico con corrección HAC simple
    try:
        ea, eb = _align(ea, eb)
        if len(ea) < 5 or len(eb) < 5:
            return None, None
        if power == 1:
            la = np.abs(ea)
            lb = np.abs(eb)
        else:
            la = ea**2
            lb = eb**2
        d = la - lb
        T = len(d)
        dbar = float(np.mean(d))

        def acov(x, k):
            x = x - np.mean(x)
            return float(np.sum(x[:T-k] * x[k:]) / T)

        var_d = acov(d, 0)
        for k in range(1, h):
            gam = acov(d, k)
            var_d += 2 * (1 - k/h) * gam
        if var_d <= 0:
            return None, None
        DM = dbar / np.sqrt(var_d / T)
        p = 2.0 * (1.0 - student_t.cdf(abs(DM), df=max(T-1, 1)))
        return float(DM), float(p)
    except Exception:
        return None, None

def write_summary_text(validacion_xlsx_path: str, validacion_dir: str) -> None:
    """
    Genera/actualiza hoja 'summary_text' en el Excel de validación con:
    - Ranking por Direction_Accuracy y RMSE
    - Recomendación automática
    - Resultado de DM (mejor vs segundo)
    """
    xlsx = Path(validacion_xlsx_path).resolve()
    base = Path(validacion_dir).resolve()
    if not xlsx.exists():
        print(f"[WARN] No existe Excel de validación: {xlsx}")
        return

    # 1) Cargar CSVs procesados
    by_model = _load_validacion_csvs(base)
    if not by_model:
        print(f"[WARN] No se encontraron CSVs en {base}")
        return

    rows = []
    err_series: dict[str, pd.Series] = {}  # para DM
    for name, df in by_model.items():
        # inferir nombre corto del modelo (ARIMA/LSTM/PROPHET/RW) si hace falta
        model = "MODEL"
        up = name.upper()
        for tag in ["ARIMA", "LSTM", "PROPHET", "RW"]:
            if tag in up:
                model = tag
                break

                # Normaliza direcciones si existen, para evitar strings/espacios
        if "direction_pred" in df.columns:
            df["direction_pred"] = _coerce_dir(df["direction_pred"])
        if "direction_true" in df.columns:
            df["direction_true"]  = _coerce_dir(df["direction_true"])
        da = _direction_accuracy(df)
        rmse = _rmse(df.get("y_true"), df.get("y_pred")) if {"y_true", "y_pred"} <= set(df.columns) else None

        # errores para DM (yt-yp)
        if {"y_true", "y_pred"} <= set(df.columns):
            yt, yp = _align(df["y_true"], df["y_pred"])
            if len(yt) and len(yp):
                err = yt - yp
                err_series[model] = err

        rows.append({
            "model": model,
            "Direction_Accuracy(%)": round(da, 3) if da is not None else np.nan,
            "RMSE": round(rmse, 6) if rmse is not None else np.nan,
            "n": int(len(df))
        })

    metric_df = pd.DataFrame(rows).dropna(subset=["RMSE"], how="all")
    if metric_df.empty:
        print("[WARN] No hay métricas suficientes para redactar resumen.")
        return

    # ---- 2) Ranking robusto: 1) mayor Direction_Accuracy, 2) menor RMSE ----
    # Copias y resets para garantizar misma longitud y evitar índices raros
    metric_df = metric_df.reset_index(drop=True).copy()

    da_s = metric_df["Direction_Accuracy(%)"].astype(float).fillna(-1e9).reset_index(drop=True)
    rm_s = metric_df["RMSE"].astype(float).fillna(1e9).reset_index(drop=True)

    # Tupla por fila, con la MISMA longitud que metric_df
    metric_df["__rank"] = list(zip(-da_s, rm_s))

    # Orden final (coherente con tu criterio)
    metric_df = metric_df.sort_values(
        by=["Direction_Accuracy(%)", "RMSE"],
        ascending=[False, True]
    ).reset_index(drop=True)

    best = metric_df.iloc[0]
    best_model = str(best["model"])

    # ---- 3) DM entre mejor y segundo (si existe) ----
    dm_text = "Sin comparación DM (un solo modelo válido)."
    if len(metric_df) >= 2:
        second = metric_df.iloc[1]
        a, b = best_model, str(second["model"])
        DM, p = (None, None)
        if a in err_series and b in err_series:
            DM, p = _dm_test(err_series[a], err_series[b], h=1, power=2)
        if DM is not None and p is not None:
            sig = "significativa" if p <= 0.10 else "no significativa"
            dm_text = f"Prueba DM entre {a} (mejor) y {b}: DM={DM:.3f}, p={p:.3f} ({sig} a 10%)."
        else:
            dm_text = f"Prueba DM entre {a} y {b}: no disponible (series insuficientes)."

    # ---- 4) Redacción de texto ----
    da_txt = f"{best['Direction_Accuracy(%)']:.2f}%" if pd.notna(best['Direction_Accuracy(%)']) else "N/D"
    rmse_txt = f"{best['RMSE']:.6f}" if pd.notna(best['RMSE']) else "N/D"

    recomendacion = (
        f"Recomendación: **{best_model}**. Motivo: mayor *Direction Accuracy* ({da_txt}) "
        f"y RMSE competitivo ({rmse_txt}). {dm_text}"
    )

    bullets = [
        "• Métrica primaria: *Direction Accuracy* (acierto direccional).",
        "• Desempate / refuerzo: *RMSE* (error cuadrático medio).",
        "• Contraste estadístico: *Diebold–Mariano (DM)* entre el mejor y el segundo.",
        "• Si DM p ≤ 0.10, la ventaja del mejor se considera estadísticamente significativa.",
    ]

    # ---- 5) Escribir hoja 'summary_text' (no tocar otras hojas) ----
    try:
        with pd.ExcelWriter(xlsx, mode="a", engine="openpyxl", if_sheet_exists="replace") as wr:
            # Hoja de texto
            resumen_df = pd.DataFrame({
                "summary_text": [
                    "RESUMEN EJECUTIVO (validación + backtest)",
                    "",
                    f"Mejor modelo: {best_model}",
                    f"Direction Accuracy (mejor): {da_txt}",
                    f"RMSE (mejor): {rmse_txt}",
                    dm_text,
                    "",
                    recomendacion,
                    "",
                    "Criterios:",
                    *bullets
                ]
            })
            resumen_df.to_excel(wr, index=False, sheet_name="summary_text")

            # Hoja con ranking numérico (por transparencia)
            metric_df.drop(columns=["__rank"], errors="ignore").to_excel(
                wr, index=False, sheet_name="summary_ranking"
            )

        print(f"[INFO] Hoja 'summary_text' escrita en: {xlsx}")
    except Exception as e:
        print(f"[WARN] No se pudo escribir 'summary_text' en {xlsx}: {e}")


if __name__ == "__main__":
    main()
